from __future__ import annotations

import csv
import hashlib
import io
import re
import secrets
import unicodedata
import uuid
import zipfile
from xml.etree import ElementTree
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

from sqlalchemy import BigInteger, Boolean, Column, DateTime, Float, MetaData, Table, Text, inspect, text
from sqlalchemy.engine import URL
from sqlalchemy.orm import Session
from sqlalchemy.schema import CreateSchema, DropSchema
from openpyxl import load_workbook

from app.connectors.base import ColumnMetadata, ConnectorMetadata, TableMetadata
from app.core.config import get_settings
from app.core.data_safety import SENSITIVE_COMMENT_MARKER, is_sensitive_column
from app.core.security import encrypt_secret
from app.file_multimodal.contracts import AttachmentKind, ParsedAttachment, TableData
from app.file_multimodal.parsers import FileParseError, PromptInjectionDetected, parse_attachment
from app.file_multimodal.security import reject_prompt_injection
from app.models import DataSource, DataSourceImport
from app.services.datasources import store_metadata


class SpreadsheetImportError(ValueError):
    def __init__(self, code: str, *, status_code: int = 400):
        super().__init__(code)
        self.code = code
        self.status_code = status_code


_SUPPORTED_EXTENSIONS = {".xlsx", ".csv"}
_CANONICAL_MIME = {
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".csv": "text/csv",
}
_ACCEPTED_MIME = {
    ".xlsx": {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
        "application/zip",
        "application/octet-stream",
    },
    ".csv": {
        "text/csv", "application/csv", "text/plain", "application/vnd.ms-excel", "application/octet-stream",
    },
}
_PROHIBITED_XLSX_PARTS = (
    "vbaproject", "externallinks/", "connections.xml", "embeddings/", "oleobjects/", "activex/",
)
_DANGEROUS_CELL = re.compile(r"^[\s\t\r\n]*[=+@-]")
_RESERVED_IDENTIFIERS = {
    "all", "and", "as", "by", "case", "create", "delete", "drop", "else", "end", "from", "group",
    "having", "insert", "into", "join", "limit", "not", "null", "on", "or", "order", "select", "table",
    "then", "union", "update", "user", "when", "where", "with",
}
_MAX_OFFICE_MEMBERS = 10_000
_MAX_OFFICE_UNCOMPRESSED_BYTES = 128 * 1024 * 1024
_MAX_OFFICE_COMPRESSION_RATIO = 200


def _xml_local_name(value: str) -> str:
    return value.rsplit("}", 1)[-1].rsplit(":", 1)[-1].lower()


def _worksheet_contains_formula(payload: bytes) -> bool:
    return any(
        _xml_local_name(element.tag) == "f"
        for _event, element in ElementTree.iterparse(io.BytesIO(payload), events=("start",))
    )


def _relationships_contain_external_target(payload: bytes) -> bool:
    for _event, element in ElementTree.iterparse(io.BytesIO(payload), events=("start",)):
        if _xml_local_name(element.tag) != "relationship":
            continue
        attributes = {_xml_local_name(name): value for name, value in element.attrib.items()}
        if str(attributes.get("targetmode", "")).strip().lower() == "external":
            return True
    return False


def _canonical_mime(filename: str, declared_mime: str) -> str:
    extension = Path(filename).suffix.lower()
    if extension not in _SUPPORTED_EXTENSIONS:
        raise SpreadsheetImportError("SPREADSHEET_FORMAT_NOT_SUPPORTED")
    normalized = (declared_mime or "application/octet-stream").split(";", 1)[0].strip().lower()
    if normalized not in _ACCEPTED_MIME[extension]:
        raise SpreadsheetImportError("SPREADSHEET_MIME_EXTENSION_MISMATCH")
    return _CANONICAL_MIME[extension]


def _reject_active_workbook_content(filename: str, data: bytes) -> None:
    if Path(filename).suffix.lower() != ".xlsx":
        return
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            members = archive.infolist()
            total_uncompressed = sum(item.file_size for item in members)
            total_compressed = sum(item.compress_size for item in members)
            if (
                len(members) > _MAX_OFFICE_MEMBERS
                or total_uncompressed > _MAX_OFFICE_UNCOMPRESSED_BYTES
                or any(item.file_size > _MAX_OFFICE_UNCOMPRESSED_BYTES for item in members)
                or (
                    total_uncompressed > 0
                    and total_uncompressed / max(1, total_compressed) > _MAX_OFFICE_COMPRESSION_RATIO
                )
            ):
                raise SpreadsheetImportError("SPREADSHEET_ARCHIVE_LIMIT_EXCEEDED")
            lowered_names = [item.filename.lower() for item in members]
            if any(part in name for name in lowered_names for part in _PROHIBITED_XLSX_PARTS):
                raise SpreadsheetImportError("SPREADSHEET_ACTIVE_CONTENT_REJECTED")
            relationships = [item for item in members if item.filename.lower().endswith(".rels")]
            for item in relationships:
                if _relationships_contain_external_target(archive.read(item)):
                    raise SpreadsheetImportError("SPREADSHEET_ACTIVE_CONTENT_REJECTED")
            worksheets = [
                item
                for item in members
                if item.filename.lower().startswith("xl/worksheets/")
                and item.filename.lower().endswith(".xml")
            ]
            if len(worksheets) > get_settings().spreadsheet_import_max_sheets:
                raise SpreadsheetImportError("SPREADSHEET_SHEET_LIMIT_EXCEEDED")
            for item in worksheets:
                if _worksheet_contains_formula(archive.read(item)):
                    raise SpreadsheetImportError("SPREADSHEET_FORMULA_REJECTED")
    except SpreadsheetImportError:
        raise
    except (zipfile.BadZipFile, OSError, ElementTree.ParseError) as exc:
        raise SpreadsheetImportError("SPREADSHEET_ARCHIVE_INVALID") from exc


def _preflight_shape(filename: str, data: bytes) -> None:
    """Reject oversized logical tables before pandas expands them in memory."""

    settings = get_settings()
    if Path(filename).suffix.lower() == ".csv":
        if b"\x00" in data[:8192]:
            raise SpreadsheetImportError("SPREADSHEET_BINARY_TEXT_FILE")
        try:
            reader = csv.reader(io.StringIO(data.decode("utf-8-sig")))
            header = next(reader, [])
            max_width = len(header)
            row_count = 0
            for row in reader:
                row_count += 1
                max_width = max(max_width, len(row))
                if row_count > settings.spreadsheet_import_max_rows:
                    raise SpreadsheetImportError("SPREADSHEET_ROW_LIMIT_EXCEEDED")
                if max_width > settings.spreadsheet_import_max_columns:
                    raise SpreadsheetImportError("SPREADSHEET_COLUMN_LIMIT_EXCEEDED")
                if row_count * max_width > settings.spreadsheet_import_max_cells:
                    raise SpreadsheetImportError("SPREADSHEET_CELL_LIMIT_EXCEEDED")
        except UnicodeDecodeError as exc:
            raise SpreadsheetImportError("SPREADSHEET_TEXT_MUST_BE_UTF8") from exc
        if max_width > settings.spreadsheet_import_max_columns:
            raise SpreadsheetImportError("SPREADSHEET_COLUMN_LIMIT_EXCEEDED")
        return

    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=False, keep_links=False)
    except Exception as exc:
        raise SpreadsheetImportError("SPREADSHEET_ARCHIVE_INVALID") from exc
    try:
        if len(workbook.worksheets) > settings.spreadsheet_import_max_sheets:
            raise SpreadsheetImportError("SPREADSHEET_SHEET_LIMIT_EXCEEDED")
        total_rows = 0
        total_cells = 0
        for worksheet in workbook.worksheets:
            # XLSX <dimension> metadata is attacker-controlled and may claim
            # A1:A1 while the worksheet XML contains far more cells. Reset it
            # and count the actual streamed rows before pandas allocates a
            # DataFrame.
            worksheet.reset_dimensions()
            sheet_rows = 0
            columns = 0
            for row in worksheet.iter_rows(values_only=True):
                sheet_rows += 1
                columns = max(columns, len(row))
                rows = max(0, sheet_rows - 1)
                if columns > settings.spreadsheet_import_max_columns:
                    raise SpreadsheetImportError("SPREADSHEET_COLUMN_LIMIT_EXCEEDED")
                if total_rows + rows > settings.spreadsheet_import_max_rows:
                    raise SpreadsheetImportError("SPREADSHEET_ROW_LIMIT_EXCEEDED")
                if total_cells + rows * columns > settings.spreadsheet_import_max_cells:
                    raise SpreadsheetImportError("SPREADSHEET_CELL_LIMIT_EXCEEDED")
            rows = max(0, sheet_rows - 1)
            total_rows += rows
            total_cells += rows * columns
    finally:
        workbook.close()


def _safe_identifier(value: str, *, fallback: str) -> str:
    normalized = unicodedata.normalize("NFKC", str(value)).strip()
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    converted = "".join(
        character if (character.isalnum() or character == "_") else "_"
        for character in ascii_value
    )
    converted = re.sub(r"_+", "_", converted).strip("_").lower() or fallback
    if converted[0].isdigit():
        converted = f"n_{converted}"
    if converted in _RESERVED_IDENTIFIERS:
        converted = f"col_{converted}"
    # PostgreSQL's 63-byte identifier limit is byte-based. Keep generated
    # names ASCII and add a stable suffix whenever transliteration discarded
    # information or truncation is required, preventing silent collisions.
    lossy = ascii_value != normalized or len(converted.encode("ascii")) > 56
    if lossy:
        suffix = hashlib.sha256(normalized.encode("utf-8")).hexdigest()[:10]
        converted = f"{converted[:45].rstrip('_') or fallback}_{suffix}"
    return converted[:56]


def _unique_identifier(value: str, *, fallback: str, used: set[str]) -> str:
    base = _safe_identifier(value, fallback=fallback)
    candidate = base
    suffix = 2
    while candidate in used:
        tail = f"_{suffix}"
        candidate = f"{base[:56 - len(tail)]}{tail}"
        suffix += 1
    used.add(candidate)
    return candidate


def _safe_json_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def _reject_formula_like_values(parsed: ParsedAttachment) -> None:
    for table in parsed.tables:
        for row in table.rows:
            for value in row.values():
                if isinstance(value, str) and _DANGEROUS_CELL.search(value):
                    raise SpreadsheetImportError("SPREADSHEET_FORMULA_REJECTED")


def _parse_and_validate(filename: str, declared_mime: str, data: bytes) -> ParsedAttachment:
    settings = get_settings()
    if not data:
        raise SpreadsheetImportError("SPREADSHEET_EMPTY_FILE")
    if len(data) > settings.spreadsheet_import_max_bytes:
        raise SpreadsheetImportError("SPREADSHEET_FILE_SIZE_LIMIT_EXCEEDED", status_code=413)
    canonical_mime = _canonical_mime(filename, declared_mime)
    _reject_active_workbook_content(filename, data)
    _preflight_shape(filename, data)
    try:
        parsed = parse_attachment(
            filename,
            canonical_mime,
            data,
            max_rows=settings.spreadsheet_import_max_rows,
        )
    except PromptInjectionDetected as exc:
        raise SpreadsheetImportError("SPREADSHEET_UNSAFE_CELL_CONTENT") from exc
    except FileParseError as exc:
        raise SpreadsheetImportError(str(exc)) from exc
    if parsed.kind != AttachmentKind.STRUCTURED or not parsed.tables:
        raise SpreadsheetImportError("SPREADSHEET_TABLE_DATA_REQUIRED")
    if len(parsed.tables) > settings.spreadsheet_import_max_sheets:
        raise SpreadsheetImportError("SPREADSHEET_SHEET_LIMIT_EXCEEDED")
    try:
        reject_prompt_injection(
            value
            for table in parsed.tables
            for value in (table.name, *table.columns)
        )
    except PromptInjectionDetected as exc:
        raise SpreadsheetImportError("SPREADSHEET_UNSAFE_METADATA") from exc
    for table in parsed.tables:
        if not table.columns:
            raise SpreadsheetImportError("SPREADSHEET_COLUMNS_REQUIRED")
        if len(table.columns) > settings.spreadsheet_import_max_columns:
            raise SpreadsheetImportError("SPREADSHEET_COLUMN_LIMIT_EXCEEDED")
    _reject_formula_like_values(parsed)
    return parsed


def _datetime_value(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if not isinstance(value, str) or not re.match(r"^\d{4}-\d{1,2}-\d{1,2}(?:[T ]|$)", value.strip()):
        return None
    try:
        parsed = datetime.fromisoformat(value.strip().replace("Z", "+00:00"))
        return parsed
    except ValueError:
        return None


def _infer_column(values: list[Any]) -> tuple[Any, str, Callable[[Any], Any]]:
    present = [value for value in values if value is not None]
    if not present:
        return Text(), "TEXT", lambda value: None if value is None else str(value)
    if all(isinstance(value, bool) for value in present):
        return Boolean(), "BOOLEAN", lambda value: value
    if all(isinstance(value, int) and not isinstance(value, bool) for value in present):
        return BigInteger(), "BIGINT", lambda value: value
    if all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in present):
        return Float(), "DOUBLE PRECISION", lambda value: None if value is None else float(value)
    parsed_dates = [_datetime_value(value) for value in present]
    if all(value is not None for value in parsed_dates):
        return DateTime(timezone=True), "TIMESTAMP", lambda value: None if value is None else _datetime_value(value)
    return Text(), "TEXT", lambda value: None if value is None else str(value)


def _normalized_tables(parsed: ParsedAttachment) -> list[dict[str, Any]]:
    used_tables: set[str] = set()
    result: list[dict[str, Any]] = []
    for table_index, source in enumerate(parsed.tables, start=1):
        table_name = _unique_identifier(
            f"sheet_{source.name}", fallback=f"sheet_{table_index}", used=used_tables,
        )
        used_columns: set[str] = set()
        columns: list[dict[str, Any]] = []
        for column_index, original_name in enumerate(source.columns, start=1):
            name = _unique_identifier(original_name, fallback=f"column_{column_index}", used=used_columns)
            values = [row.get(original_name) for row in source.rows]
            sql_type, type_name, converter = _infer_column(values)
            sensitive = is_sensitive_column(name) or is_sensitive_column(original_name)
            columns.append({
                "name": name,
                "source_name": original_name,
                "sql_type": sql_type,
                "type_name": type_name,
                "converter": converter,
                "nullable": any(value is None for value in values),
                "sensitive": sensitive,
                "sample_values": (
                    [] if sensitive
                    else [_safe_json_value(value) for value in values if value is not None][:5]
                ),
            })
        result.append({
            "source_name": source.name,
            "name": table_name,
            "row_count": source.row_count,
            "columns": columns,
            # Keep the parser's immutable row representation by reference.
            # Conversion is performed only for the preview slice or a bounded
            # database insert batch, avoiding a second full workbook copy.
            "source_rows": source.rows,
        })
    return result


def _converted_row(
    table: dict[str, Any], source_row: Any, *, redact_sensitive: bool = False,
) -> dict[str, Any]:
    return {
        column["name"]: (
            None
            if redact_sensitive and column["sensitive"]
            else column["converter"](source_row.get(column["source_name"]))
        )
        for column in table["columns"]
    }


def _preview_payload(parsed: ParsedAttachment, tables: list[dict[str, Any]], *, file_size: int) -> dict[str, Any]:
    settings = get_settings()
    return {
        "filename": parsed.filename,
        "file_sha256": parsed.file_sha256,
        "file_size_bytes": file_size,
        "format": parsed.extension.lstrip("."),
        "sheet_count": len(tables),
        "row_count": sum(item["row_count"] for item in tables),
        "column_count": sum(len(item["columns"]) for item in tables),
        "limits": {
            "max_bytes": settings.spreadsheet_import_max_bytes,
            "max_rows": settings.spreadsheet_import_max_rows,
            "max_columns_per_sheet": settings.spreadsheet_import_max_columns,
            "max_sheets": settings.spreadsheet_import_max_sheets,
            "max_cells": settings.spreadsheet_import_max_cells,
        },
        "sheets": [
            {
                "source_name": item["source_name"],
                "table_name": item["name"],
                "row_count": item["row_count"],
                "columns": [
                    {
                        "source_name": column["source_name"],
                        "name": column["name"],
                        "data_type": column["type_name"],
                        "nullable": column["nullable"],
                    }
                    for column in item["columns"]
                ],
                "preview_rows": [
                    {key: _safe_json_value(value) for key, value in row.items()}
                    for row in (
                        _converted_row(item, source_row, redact_sensitive=True)
                        for source_row in item["source_rows"][:settings.spreadsheet_import_preview_rows]
                    )
                ],
            }
            for item in tables
        ],
    }


def spreadsheet_preview(filename: str, declared_mime: str, data: bytes) -> dict[str, Any]:
    parsed = _parse_and_validate(filename, declared_mime, data)
    return _preview_payload(parsed, _normalized_tables(parsed), file_size=len(data))


def _connection_details(db: Session, *, datasource_id: str) -> tuple[URL, str, str]:
    bind = db.get_bind()
    url = getattr(bind, "url", None)
    if url is None:
        raise SpreadsheetImportError("SPREADSHEET_STORAGE_DATABASE_UNAVAILABLE", status_code=503)
    if url.get_backend_name() == "postgresql":
        suffix = datasource_id.replace("-", "")[:12]
        return url, f"chatbi_excel_{suffix}", secrets.token_urlsafe(32)
    if url.get_backend_name() == "sqlite":
        # SQLite is used by isolated API tests only. Production deployments are
        # required to materialize into the configured local PostgreSQL database.
        return url, "managed-readonly", "managed-test-only"
    raise SpreadsheetImportError("SPREADSHEET_STORAGE_REQUIRES_POSTGRESQL", status_code=503)


def import_spreadsheet(
    db: Session,
    *,
    workspace_id: str,
    name: str,
    filename: str,
    declared_mime: str,
    data: bytes,
) -> tuple[DataSource, dict[str, Any]]:
    display_name = name.strip()
    if not display_name or len(display_name) > 255:
        raise SpreadsheetImportError("SPREADSHEET_DATASOURCE_NAME_INVALID")
    parsed = _parse_and_validate(filename, declared_mime, data)
    normalized = _normalized_tables(parsed)
    preview = _preview_payload(parsed, normalized, file_size=len(data))
    datasource_id = str(uuid.uuid4())
    url, reader_username, reader_password = _connection_details(db, datasource_id=datasource_id)
    storage_schema = f"excel_{datasource_id.replace('-', '')[:12]}"
    is_postgres = url.get_backend_name() == "postgresql"
    metadata_schema = storage_schema if is_postgres else "main"
    datasource = DataSource(
        id=datasource_id,
        workspace_id=workspace_id,
        name=display_name,
        type="excel",
        host=url.host or "backend-managed",
        port=url.port or (5432 if is_postgres else 1),
        database=url.database or "chatbi_v2",
        username=reader_username,
        password_encrypted=encrypt_secret(reader_password),
        ssl=False,
        schema=storage_schema if is_postgres else None,
        status="PENDING",
    )
    db.add(datasource)
    db.flush()

    try:
        if is_postgres:
            db.execute(CreateSchema(storage_schema))
        sql_metadata = MetaData()
        connector_metadata = ConnectorMetadata(schemas=[metadata_schema])
        sheet_metadata: list[dict[str, Any]] = []
        for item in normalized:
            physical_name = item["name"] if is_postgres else f"{storage_schema}__{item['name']}"
            sql_table = Table(
                physical_name,
                sql_metadata,
                *[Column(column["name"], column["sql_type"], nullable=column["nullable"]) for column in item["columns"]],
                schema=storage_schema if is_postgres else None,
            )
            sql_table.create(bind=db.connection())
            source_rows = item["source_rows"]
            for start in range(0, len(source_rows), 1000):
                batch = [
                    _converted_row(item, source_row)
                    for source_row in source_rows[start:start + 1000]
                ]
                if batch:
                    db.execute(sql_table.insert(), batch)
            catalog_name = item["name"] if is_postgres else physical_name
            connector_metadata.tables.append(TableMetadata(
                schema=metadata_schema,
                name=catalog_name,
                comment=f"Imported from spreadsheet sheet: {item['source_name']}",
                columns=[
                    ColumnMetadata(
                        name=column["name"],
                        data_type=column["type_name"],
                        nullable=column["nullable"],
                        comment=(
                            " ".join(filter(None, [
                                SENSITIVE_COMMENT_MARKER if column["sensitive"] else None,
                                (
                                    f"Original column: {column['source_name']}"
                                    if column["source_name"] != column["name"] else None
                                ),
                            ])) or None
                        ),
                        sample_values=column["sample_values"],
                    )
                    for column in item["columns"]
                ],
            ))
            sheet_metadata.append({
                "source_name": item["source_name"],
                "table_name": catalog_name,
                "storage_table": physical_name,
                "row_count": item["row_count"],
                "columns": [
                    {
                        "source_name": column["source_name"],
                        "name": column["name"],
                        "data_type": column["type_name"],
                    }
                    for column in item["columns"]
                ],
            })
        if is_postgres:
            helper = db.scalar(text(
                "SELECT to_regprocedure('chatbi_admin.provision_excel_reader(text,text,text)')"
            ))
            if helper is None:
                raise SpreadsheetImportError(
                    "SPREADSHEET_READER_PROVISIONER_UNAVAILABLE", status_code=503,
                )
            db.execute(
                text("SELECT chatbi_admin.provision_excel_reader(:role, :password, :schema)"),
                {"role": reader_username, "password": reader_password, "schema": storage_schema},
            )
        import_record = DataSourceImport(
            datasource_id=datasource.id,
            original_filename=parsed.filename,
            file_sha256=parsed.file_sha256,
            media_type=parsed.mime_type,
            file_size_bytes=len(data),
            storage_schema=storage_schema,
            row_count=sum(item["row_count"] for item in normalized),
            column_count=sum(len(item["columns"]) for item in normalized),
            sheet_metadata=sheet_metadata,
            status="READY",
        )
        db.add(import_record)
        counts = store_metadata(db, datasource, connector_metadata, commit=False)
    except Exception:
        db.rollback()
        raise

    preview["datasource_id"] = datasource.id
    preview["storage_schema"] = storage_schema
    preview["catalog"] = counts
    return datasource, preview


def delete_managed_datasource(db: Session, datasource: DataSource) -> None:
    import_record = db.get(DataSourceImport, datasource.spreadsheet_import.id) if datasource.spreadsheet_import else None
    if import_record is not None:
        expected_schema = f"excel_{datasource.id.replace('-', '')[:12]}"
        if import_record.storage_schema != expected_schema or not re.fullmatch(r"excel_[0-9a-f]{12}", expected_schema):
            raise SpreadsheetImportError("SPREADSHEET_STORAGE_TARGET_INVALID")
        bind = db.get_bind()
        if bind.dialect.name == "postgresql":
            expected_role = f"chatbi_excel_{datasource.id.replace('-', '')[:12]}"
            if datasource.username == expected_role:
                helper = db.scalar(text(
                    "SELECT to_regprocedure('chatbi_admin.drop_excel_reader(text,text)')"
                ))
                if helper is None:
                    raise SpreadsheetImportError(
                        "SPREADSHEET_READER_PROVISIONER_UNAVAILABLE", status_code=503,
                    )
                db.execute(
                    text("SELECT chatbi_admin.drop_excel_reader(:role, :schema)"),
                    {"role": expected_role, "schema": expected_schema},
                )
            elif datasource.username != get_settings().demo_postgres_username.strip():
                raise SpreadsheetImportError("SPREADSHEET_STORAGE_TARGET_INVALID")
            db.execute(DropSchema(expected_schema, cascade=True, if_exists=True))
        elif bind.dialect.name == "sqlite":
            inspector = inspect(bind)
            allowed_tables = {
                item.get("storage_table") for item in import_record.sheet_metadata if isinstance(item, dict)
            }
            for table_name in allowed_tables:
                if not isinstance(table_name, str) or not table_name.startswith(f"{expected_schema}__"):
                    raise SpreadsheetImportError("SPREADSHEET_STORAGE_TARGET_INVALID")
                if table_name in inspector.get_table_names():
                    Table(table_name, MetaData()).drop(bind=db.connection())
    db.delete(datasource)
    db.commit()


__all__ = [
    "SpreadsheetImportError",
    "delete_managed_datasource",
    "import_spreadsheet",
    "spreadsheet_preview",
]
