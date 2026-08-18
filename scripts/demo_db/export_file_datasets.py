from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import Workbook
from sqlalchemy import create_engine, text
from sqlalchemy.engine import URL

from _common import atomic_json, canonical_hash, connection_kwargs, file_hash, load_env, validate_schema


SALES_COLUMNS = [
    "tenant_id", "order_id", "order_date", "customer_id", "product_id", "region_id",
    "quantity", "unit_price", "discount_rate", "gross_amount", "net_amount", "cost_amount",
    "profit_amount", "order_status", "refund_amount", "sales_channel",
]

FIXED_DOCUMENT_TIME = datetime(2026, 8, 18, tzinfo=timezone.utc)


def save_deterministic_workbook(workbook: Workbook, path: Path) -> None:
    workbook.properties.created = FIXED_DOCUMENT_TIME.replace(tzinfo=None)
    workbook.properties.modified = FIXED_DOCUMENT_TIME.replace(tzinfo=None)
    workbook.save(path)
    temporary = path.with_suffix(path.suffix + ".normalized")
    with ZipFile(path, "r") as source, ZipFile(temporary, "w", compression=ZIP_DEFLATED, compresslevel=6) as target:
        for source_info in source.infolist():
            normalized = ZipInfo(source_info.filename, date_time=(2026, 8, 18, 0, 0, 0))
            normalized.compress_type = source_info.compress_type
            normalized.external_attr = source_info.external_attr
            normalized.create_system = source_info.create_system
            target.writestr(normalized, source.read(source_info.filename))
    temporary.replace(path)


def engine_from_env(env: dict[str, str]):
    cfg = connection_kwargs(env)
    url = URL.create("postgresql+psycopg", username=cfg["user"], password=cfg["password"], host=cfg["host"], port=cfg["port"], database=cfg["dbname"])
    return create_engine(url)


def write_parquet(engine, schema: str, path: Path, rows: int) -> int:
    query = text(f'SELECT {", ".join(SALES_COLUMNS)} FROM "{schema}".fact_sales ORDER BY order_id LIMIT {rows}')
    writer = None
    total = 0
    try:
        for frame in pd.read_sql_query(query, engine, chunksize=100_000):
            table = pa.Table.from_pandas(frame, preserve_index=False)
            writer = writer or pq.ParquetWriter(path, table.schema, compression="zstd")
            writer.write_table(table)
            total += len(frame)
    finally:
        if writer:
            writer.close()
    return total


def write_csv(engine, schema: str, path: Path, rows: int) -> int:
    query = text(f'SELECT {", ".join(SALES_COLUMNS)} FROM "{schema}".fact_sales ORDER BY order_id LIMIT {rows}')
    total = 0
    first = True
    for frame in pd.read_sql_query(query, engine, chunksize=50_000):
        frame.to_csv(path, mode="w" if first else "a", header=first, index=False, encoding="utf-8")
        first = False
        total += len(frame)
    return total


def write_query_sheet(workbook: Workbook, title: str, engine, query: str, limit: int | None = None) -> int:
    sheet = workbook.create_sheet(title)
    total = 0
    for frame in pd.read_sql_query(text(query), engine, chunksize=10_000):
        if total == 0:
            sheet.append(list(frame.columns))
        for row in frame.itertuples(index=False, name=None):
            sheet.append(list(row))
        total += len(frame)
        if limit and total >= limit:
            break
    return total


def write_sales_workbook(engine, schema: str, path: Path) -> dict[str, int]:
    workbook = Workbook(write_only=True)
    sales = write_query_sheet(
        workbook,
        "sales",
        engine,
        f'SELECT {", ".join(SALES_COLUMNS)} FROM "{schema}".fact_sales ORDER BY order_id LIMIT 50000',
    )
    monthly = write_query_sheet(
        workbook,
        "monthly_summary",
        engine,
        f'SELECT * FROM "{schema}".agg_monthly_sales ORDER BY tenant_id, month_start',
    )
    save_deterministic_workbook(workbook, path)
    return {"sales": sales, "monthly_summary": monthly}


def write_customer_workbook(engine, schema: str, path: Path) -> int:
    workbook = Workbook(write_only=True)
    count = write_query_sheet(
        workbook,
        "customers",
        engine,
        f'SELECT customer_id, customer_code, customer_name, customer_tier, industry, registered_region_id, registered_at, email, phone FROM "{schema}".dim_customer ORDER BY customer_id LIMIT 100000',
    )
    save_deterministic_workbook(workbook, path)
    return count


def build_file_golden(engine, schema: str, data_signature: str) -> dict:
    cases = []
    for tenant in range(1, 11):
        query = text(
            f'''SELECT round(sum(net_amount - refund_amount)::numeric, 2) AS net_sales
                FROM "{schema}".fact_sales
                WHERE tenant_id=:tenant AND order_status IN ('VALID','PARTIAL_REFUND','REFUNDED')'''
        )
        with engine.connect() as conn:
            value = conn.execute(query, {"tenant": tenant}).scalar()
        expected = {"net_sales": str(value)}
        cases.append({
            "case_id": f"FILE-{tenant:02d}",
            "question": f"租户 {tenant} 的净销售额是多少？",
            "dataset": "sales_sample.parquet",
            "expected": expected,
            "result_signature": canonical_hash(expected),
        })
    return {"count": len(cases), "data_signature": data_signature, "cases": cases}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--env-file", type=Path, default=Path(".env"))
    parser.add_argument("--schema", default="chatbi_benchmark_v21")
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--data-signature", required=True)
    args = parser.parse_args()
    schema = validate_schema(args.schema)
    args.output_dir.mkdir(parents=True, exist_ok=True)
    engine = engine_from_env(load_env(args.env_file))

    files: dict[str, dict] = {}
    parquet_path = args.output_dir / "sales_sample.parquet"
    files[parquet_path.name] = {"rows": write_parquet(engine, schema, parquet_path, 1_000_000)}
    csv_path = args.output_dir / "sales_sample.csv"
    files[csv_path.name] = {"rows": write_csv(engine, schema, csv_path, 200_000)}
    sales_xlsx = args.output_dir / "sales_workbook.xlsx"
    files[sales_xlsx.name] = {"sheets": write_sales_workbook(engine, schema, sales_xlsx)}
    customer_xlsx = args.output_dir / "customer_profile.xlsx"
    files[customer_xlsx.name] = {"rows": write_customer_workbook(engine, schema, customer_xlsx)}
    golden_path = args.output_dir / "file_golden.json"
    golden = build_file_golden(engine, schema, args.data_signature)
    atomic_json(golden_path, golden)
    files[golden_path.name] = {"cases": golden["count"]}

    for name, metadata in files.items():
        path = args.output_dir / name
        metadata.update({"bytes": path.stat().st_size, "sha256": file_hash(path)})
    atomic_json(args.output_dir / "file_datasets_manifest.json", {"schema": schema, "data_signature": args.data_signature, "files": files})
    print(json.dumps(files, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
